import pandas as pd
import os
import re
from openpyxl import load_workbook

input_file = 'MSL.xlsx'
sheet_name = 'MSL'
output_dir = 'converted_files'

os.makedirs(output_dir, exist_ok=True)

# --- STEP 1: Load data normally with pandas ---
df = pd.read_excel(input_file, sheet_name=sheet_name)

# Drop unwanted index column if present
if 'Unnamed: 0' in df.columns:
    df = df.drop(columns=['Unnamed: 0'])

# Clean column names for XML tags (keep consistent)
df.columns = [re.sub(r'\W+', '_', str(col)) for col in df.columns]

# --- STEP 2: Extract hyperlinks from Column AB (index 28 in Excel, 0-based index 27) ---
wb = load_workbook(input_file, data_only=True)
ws = wb[sheet_name]

# Find the correct column (AB = 28th column, index 27)
col_idx = 28

hyperlinks = []
for row in range(2, ws.max_row + 1):  # skip header row
    cell = ws.cell(row=row, column=col_idx)
    if cell.hyperlink:
        hyperlinks.append(cell.hyperlink.target)
    else:
        hyperlinks.append(None)

# Add the hyperlinks as a new column in df
df["from_url"] = hyperlinks

# --- STEP 3: Save converted formats ---
df.to_csv(os.path.join(output_dir, f'{sheet_name}.tsv'), sep='\t', index=False)
df.to_json(os.path.join(output_dir, f'{sheet_name}.json'), orient='records', indent=2)
df.to_xml(os.path.join(output_dir, f'{sheet_name}.xml'), index=False)

print(f"Conversion completed successfully. Saved to '{output_dir}/'")

# import pandas as pd
# import os
# import re

# input_file = 'MSL.xlsx'
# sheet_name = 'MSL'
# output_dir = 'converted_files'

# os.makedirs(output_dir, exist_ok=True)

# # Load the specific sheet
# df = pd.read_excel(input_file, sheet_name=sheet_name)

# # Drop unwanted index column if present
# if 'Unnamed: 0' in df.columns:
#     df = df.drop(columns=['Unnamed: 0'])

# # Clean column names for XML tags
# df.columns = [re.sub(r'\W+', '_', str(col)) for col in df.columns]

# # Export to TSV
# df.to_csv(os.path.join(output_dir, f'{sheet_name}.tsv'), sep='\t', index=False)

# # Export to JSON
# df.to_json(os.path.join(output_dir, f'{sheet_name}.json'), orient='records', indent=2)

# # Export to XML
# df.to_xml(os.path.join(output_dir, f'{sheet_name}.xml'), index=False)

# print("Conversion completed successfully.")

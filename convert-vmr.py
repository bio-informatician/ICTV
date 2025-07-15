import pandas as pd
import os
import re
from openpyxl import load_workbook

input_file = 'VMR.xlsx'
sheet_name = 'VMR MSL40'
output_dir = 'converted_files'
output_file = 'VMR'

os.makedirs(output_dir, exist_ok=True)

# --- STEP 1: Load table normally with pandas ---
df = pd.read_excel(input_file, sheet_name=sheet_name)

# Drop unwanted index column if present
if 'Unnamed: 0' in df.columns:
    df = df.drop(columns=['Unnamed: 0'])

# Clean column names for XML tags
df.columns = [re.sub(r'\W+', '_', str(col)) for col in df.columns]

# --- STEP 2: Extract real hyperlinks from column AB ---
wb = load_workbook(input_file, data_only=False)  # keep formulas
ws = wb[sheet_name]

col_idx = 28  # AB column in Excel (1-based index)

urls = []
for row in range(2, ws.max_row + 1):  # skip header row
    cell = ws.cell(row=row, column=col_idx)
    url = None

    # 1. If Excel stored as real hyperlink object
    if cell.hyperlink:
        url = cell.hyperlink.target

    # 2. If stored as a HYPERLINK formula (e.g., =HYPERLINK("url","text"))
    elif isinstance(cell.value, str) and cell.value.startswith("=HYPERLINK"):
        match = re.search(r'HYPERLINK\("([^"]+)"', cell.value)
        if match:
            url = match.group(1)

    urls.append(url)

# Add as a new column
df["from_url"] = urls

# --- STEP 3: Save all formats ---
df.to_csv(os.path.join(output_dir, f'{output_file}.tsv'), sep='\t', index=False)
df.to_json(os.path.join(output_dir, f'{output_file}.json'), orient='records', indent=2)
df.to_xml(os.path.join(output_dir, f'{output_file}.xml'), index=False)

print(f"✅ Conversion completed successfully. Saved as {output_file}.* in '{output_dir}/'")

# import pandas as pd
# import os
# import re

# input_file = 'VMR.xlsx'
# sheet_name = 'VMR MSL40'
# output_dir = 'converted_files'
# output_file = 'VMR'

# os.makedirs(output_dir, exist_ok=True)

# # Load the specific sheet
# df = pd.read_excel(input_file, sheet_name=sheet_name)

# # Drop unwanted index column if present
# if 'Unnamed: 0' in df.columns:
#     df = df.drop(columns=['Unnamed: 0'])

# # Clean column names for XML tags
# df.columns = [re.sub(r'\W+', '_', str(col)) for col in df.columns]

# # Export to TSV
# df.to_csv(os.path.join(output_dir, f'{output_file}.tsv'), sep='\t', index=False)

# # Export to JSON
# df.to_json(os.path.join(output_dir, f'{output_file}.json'), orient='records', indent=2)

# # Export to XML
# df.to_xml(os.path.join(output_dir, f'{output_file}.xml'), index=False)

# print("Conversion completed successfully.")
